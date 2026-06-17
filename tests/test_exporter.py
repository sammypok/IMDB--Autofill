"""
test_exporter.py — Unit tests for backend/exporter.py.

Tests verify that build_csv_bytes and build_xlsx_bytes:
  - Produce headers that exactly match IMDB_FIELDS (including two-space PACKAGING  TYPE)
  - Write the correct number of data rows
  - Use empty string (not None) for missing fields
  - Handle empty input without crashing
  - Produce valid UTF-8 bytes
"""

import csv
import io

import openpyxl
import pytest

from backend.models import IMDB_FIELDS
from backend.exporter import build_csv_bytes, build_xlsx_bytes

# ---------------------------------------------------------------------------
# Fixtures / shared data
# ---------------------------------------------------------------------------

SAMPLE_RESULT = {
    "fields": {f: f"val_{i}" for i, f in enumerate(IMDB_FIELDS)},
    "confidence": {f: "high" for f in IMDB_FIELDS},
}

EMPTY_RESULT = {
    "fields": {f: "" for f in IMDB_FIELDS},
    "confidence": {},
}


# ---------------------------------------------------------------------------
# CSV tests
# ---------------------------------------------------------------------------


def test_csv_header_matches_imdb_fields():
    """Header row must equal IMDB_FIELDS exactly (catches two-space PACKAGING  TYPE mismatch)."""
    csv_bytes = build_csv_bytes([SAMPLE_RESULT])
    rows = list(csv.reader(io.StringIO(csv_bytes.decode("utf-8"))))
    assert rows[0] == list(IMDB_FIELDS)


def test_csv_data_row_count():
    """Two results produce 3 rows total: 1 header + 2 data."""
    csv_bytes = build_csv_bytes([SAMPLE_RESULT, SAMPLE_RESULT])
    rows = list(csv.reader(io.StringIO(csv_bytes.decode("utf-8"))))
    assert len(rows) == 3


def test_csv_empty_field_is_empty_string():
    """A result with an empty field must produce '' in the CSV, not 'None'."""
    result = {
        "fields": {f: "" for f in IMDB_FIELDS},
        "confidence": {},
    }
    csv_bytes = build_csv_bytes([result])
    rows = list(csv.reader(io.StringIO(csv_bytes.decode("utf-8"))))
    data_row = rows[1]
    # No cell should contain the literal string "None"
    assert "None" not in data_row
    # Every cell should be empty string
    assert all(cell == "" for cell in data_row)


def test_csv_empty_results_returns_header_only():
    """build_csv_bytes([]) must return a single-row CSV (header only) without crashing."""
    csv_bytes = build_csv_bytes([])
    rows = list(csv.reader(io.StringIO(csv_bytes.decode("utf-8"))))
    # csv.reader may produce one empty row after a trailing newline; filter empties
    non_empty = [r for r in rows if r]
    assert len(non_empty) == 1
    assert non_empty[0] == list(IMDB_FIELDS)


def test_csv_utf8_encoding():
    """CSV bytes must be valid UTF-8 even when result contains unicode values."""
    result = {
        "fields": {f: "Café 中文" for f in IMDB_FIELDS},
        "confidence": {},
    }
    csv_bytes = build_csv_bytes([result])
    # Must decode without error
    decoded = csv_bytes.decode("utf-8")
    assert "Café" in decoded


# ---------------------------------------------------------------------------
# XLSX tests
# ---------------------------------------------------------------------------


def test_xlsx_header_matches_imdb_fields():
    """Row 1 of the XLSX must equal IMDB_FIELDS exactly."""
    xlsx_bytes = build_xlsx_bytes([SAMPLE_RESULT])
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))
    ws = wb.active
    header = [cell.value for cell in ws[1]]
    assert header == list(IMDB_FIELDS)


def test_xlsx_data_row_count():
    """Two results produce 3 rows in the worksheet: 1 header + 2 data."""
    xlsx_bytes = build_xlsx_bytes([SAMPLE_RESULT, SAMPLE_RESULT])
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))
    ws = wb.active
    assert ws.max_row == 3


def test_xlsx_empty_results_returns_header_only():
    """build_xlsx_bytes([]) must produce a single-row worksheet without crashing."""
    xlsx_bytes = build_xlsx_bytes([])
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))
    ws = wb.active
    assert ws.max_row == 1
    header = [cell.value for cell in ws[1]]
    assert header == list(IMDB_FIELDS)
